import re
from openpyxl import load_workbook

# Load the workbook
workbook = load_workbook('Workbook1.xlsx')

# Select the description sheet
description_sheet = workbook['DescriptionSheet']

# Select the keyword-comment mapping sheet
keyword_sheet = workbook['KeywordCommentSheet']

# Read the keyword-comment mapping from the sheet
keyword_mapping = {}
for row in keyword_sheet.iter_rows(min_row=2, values_only=True):
    keywords = row[0]  # Assuming keywords are in the first column
    comment = row[1]  # Assuming comment is in the second column
    keyword_mapping[keywords] = comment

# Iterate through the rows starting from the second row (assuming the first row is headers)
i = 2
probability_threshold = 0.5
for row in description_sheet.iter_rows(min_row=2, values_only=True):
    description = row[1]  # Assuming description is in the first column

    comments = row[2]  # Assuming comments is in the second column
    # Perform keyword matching and update comments
    for keywords, comment in keyword_mapping.items():
        print(comments)
        if comments is not None:
            continue
        keyword_list = keywords.split(';')  # Split multiple keywords by ';' if they are stored as a single string
        match_count = 0
        total_keywords = 0
        for keyword in keyword_list:
            pattern = re.compile(keyword.replace('*', '.*'), re.IGNORECASE)
            print("Description " + description)
            if pattern.search(description):
                print("Keyword found in description" + keyword)
                match_count += match_count + 1
            total_keywords += 1
        match_probability: float = match_count / total_keywords
        if match_probability >= probability_threshold:
            print("Probability calculated is greater than 0.5")
            comments = comment

    # Update the comments in the appropriate column
    description_sheet.cell(row=i, column=3).value = comments
    i = i + 1

# Save the updated workbook
workbook.save('Workbook1.xlsx')
